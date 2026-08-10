import openpyxl

wb = openpyxl.load_workbook(r"C:\Users\sergi\Desktop\LEO\planos_leo.xlsx", data_only=False)
print("ABAS:", wb.sheetnames)
print()

for name in wb.sheetnames:
    ws = wb[name]
    print(f"{'='*60}")
    print(f"ABA: {name}")
    print(f"{'='*60}")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=False):
        vals = []
        for cell in row:
            v = cell.value
            if v is not None:
                vals.append(f"[{cell.coordinate}] {v}")
        if vals:
            print("  |  ".join(vals))
    print()
