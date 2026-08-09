from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

DARK_BLUE = "1E3A5F"
MED_BLUE = "2563EB"
LIGHT_GRAY = "F1F5F9"
WHITE = "FFFFFF"
DARK_TEXT = "0F172A"
GREEN = "064E3B"
LIGHT_GREEN = "D1FAE5"
AMBER = "78350F"
LIGHT_AMBER = "FEF3C7"

hdr_font = Font(name="Arial", bold=True, color=WHITE, size=11)
hdr_fill = PatternFill("solid", fgColor=DARK_BLUE)
title_font = Font(name="Arial", bold=True, color=DARK_BLUE, size=14)
data_font = Font(name="Arial", size=10, color=DARK_TEXT)
blue_font = Font(name="Arial", size=10, color="0000FF")
money_fmt = 'R$ #,##0.00'
int_fmt = '#,##0'
thin_border = Border(
    left=Side(style="thin", color="CBD5E1"), right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"), bottom=Side(style="thin", color="CBD5E1"),
)
center = Alignment(horizontal="center", vertical="center")
left_align = Alignment(horizontal="left", vertical="center")

def style_hdr(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = thin_border

def style_row(ws, row, cols, alt=False):
    fill = PatternFill("solid", fgColor=LIGHT_GRAY) if alt else PatternFill("solid", fgColor=WHITE)
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = data_font
        cell.fill = fill
        cell.border = thin_border
        cell.alignment = center

# ===== ABA 1: PLANOS PF =====
ws = wb.active
ws.title = "Planos PF"
ws.sheet_properties.tabColor = DARK_BLUE
ws["A1"] = "PLANOS PESSOA FISICA \u2014 LEO Sistema de Laudos"
ws["A1"].font = title_font
ws["A2"] = "Aprovado em 11/04/2026"
ws["A2"].font = Font(name="Arial", size=9, color="64748B")

h1 = ["Plano", "Preco Mensal (R$)", "Franquia (laudos/mes)", "Excedente (R$/laudo)", "Custo por Laudo (R$)", "Publico Alvo"]
for i, h in enumerate(h1, 1):
    ws.cell(row=4, column=i, value=h)
style_hdr(ws, 4, 6)

pf = [
    ["Basic", 99.00, 100, 1.50, "Medicos baixo volume / conservadores"],
    ["Profissional", 189.99, 350, 0.75, "Maioria dos medicos (~50% esperado)"],
    ["Expert", 249.99, 500, 0.50, "Heavy users (500+ laudos/mes)"],
]
for r, p in enumerate(pf, 5):
    ws.cell(row=r, column=1, value=p[0])
    ws.cell(row=r, column=2, value=p[1])
    ws.cell(row=r, column=3, value=p[2])
    ws.cell(row=r, column=4, value=p[3])
    ws.cell(row=r, column=5).value = f"=B{r}/C{r}"
    ws.cell(row=r, column=6, value=p[4])
    ws.cell(row=r, column=2).number_format = money_fmt
    ws.cell(row=r, column=4).number_format = money_fmt
    ws.cell(row=r, column=5).number_format = money_fmt
    ws.cell(row=r, column=1).alignment = left_align
    ws.cell(row=r, column=6).alignment = left_align
    style_row(ws, r, 6, r % 2 == 0)

# Highlight Profissional
for c in range(1, 7):
    ws.cell(row=6, column=c).fill = PatternFill("solid", fgColor=LIGHT_AMBER)

for c, w in enumerate([18, 20, 24, 22, 22, 42], 1):
    ws.column_dimensions[get_column_letter(c)].width = w

# ===== ABA 2: PLANOS PJ =====
ws2 = wb.create_sheet("Planos PJ")
ws2.sheet_properties.tabColor = "4C1D95"
ws2["A1"] = "PLANOS PESSOA JURIDICA \u2014 LEO Sistema de Laudos"
ws2["A1"].font = title_font
ws2["A2"] = "A definir"
ws2["A2"].font = Font(name="Arial", size=9, color="64748B")

for i, h in enumerate(h1, 1):
    ws2.cell(row=4, column=i, value=h)
style_hdr(ws2, 4, 6)

pj = [
    ["Clinica Starter", "(a definir)", "(a definir)", "(a definir)", "Clinicas pequenas (1-3 medicos)"],
    ["Clinica Pro", "(a definir)", "(a definir)", "(a definir)", "Clinicas medias (4-10 medicos)"],
    ["Enterprise", "(a definir)", "(a definir)", "(a definir)", "Hospitais e grandes grupos"],
]
for r, p in enumerate(pj, 5):
    ws2.cell(row=r, column=1, value=p[0])
    for c in range(2, 5):
        ws2.cell(row=r, column=c, value=p[c - 1])
        ws2.cell(row=r, column=c).font = Font(name="Arial", size=10, color="94A3B8", italic=True)
    ws2.cell(row=r, column=5, value="")
    ws2.cell(row=r, column=6, value=p[4])
    ws2.cell(row=r, column=1).alignment = left_align
    ws2.cell(row=r, column=6).alignment = left_align
    style_row(ws2, r, 6, r % 2 == 0)

for c, w in enumerate([18, 20, 24, 22, 22, 42], 1):
    ws2.column_dimensions[get_column_letter(c)].width = w

# ===== ABA 3: SIMULACAO =====
ws3 = wb.create_sheet("Simulacao")
ws3.sheet_properties.tabColor = MED_BLUE
ws3["A1"] = "SIMULACAO DE CUSTO POR VOLUME"
ws3["A1"].font = title_font
ws3["A2"] = "Qual plano sai mais barato para cada volume de laudos?"
ws3["A2"].font = Font(name="Arial", size=9, color="64748B")

h3 = ["Laudos/Mes", "Custo Basic (R$)", "Custo Profissional (R$)", "Custo Expert (R$)", "Melhor Plano"]
for i, h in enumerate(h3, 1):
    ws3.cell(row=4, column=i, value=h)
style_hdr(ws3, 4, 5)

vols = [50, 100, 150, 200, 250, 300, 350, 400, 500, 750, 1000]
for r, v in enumerate(vols, 5):
    ws3.cell(row=r, column=1, value=v)
    ws3.cell(row=r, column=2).value = f"=IF(A{r}<=100,99,99+(A{r}-100)*1.5)"
    ws3.cell(row=r, column=3).value = f"=IF(A{r}<=350,189.99,189.99+(A{r}-350)*0.75)"
    ws3.cell(row=r, column=4).value = f"=IF(A{r}<=500,249.99,249.99+(A{r}-500)*0.5)"
    ws3.cell(row=r, column=5).value = f'=IF(AND(B{r}<=C{r},B{r}<=D{r}),"Basic",IF(AND(C{r}<=B{r},C{r}<=D{r}),"Profissional","Expert"))'
    ws3.cell(row=r, column=2).number_format = money_fmt
    ws3.cell(row=r, column=3).number_format = money_fmt
    ws3.cell(row=r, column=4).number_format = money_fmt
    ws3.cell(row=r, column=1).number_format = int_fmt
    style_row(ws3, r, 5, r % 2 == 0)

for c, w in enumerate([16, 22, 26, 22, 18], 1):
    ws3.column_dimensions[get_column_letter(c)].width = w

# ===== ABA 4: RECEITA ESTIMADA =====
ws4 = wb.create_sheet("Receita Estimada")
ws4.sheet_properties.tabColor = GREEN
ws4["A1"] = "PROJECAO DE RECEITA \u2014 LEO"
ws4["A1"].font = title_font
ws4["A2"] = "Distribuicao: 25% Basic / 50% Profissional / 25% Expert"
ws4["A2"].font = Font(name="Arial", size=9, color="64748B")

h4 = ["Total Usuarios", "Basic (25%)", "Profissional (50%)", "Expert (25%)", "Receita Mensal (R$)", "Receita Anual (R$)"]
for i, h in enumerate(h4, 1):
    ws4.cell(row=4, column=i, value=h)
style_hdr(ws4, 4, 6)

for r, total in enumerate([50, 100, 200, 500, 1000], 5):
    ws4.cell(row=r, column=1, value=total)
    ws4.cell(row=r, column=2).value = f"=ROUND(A{r}*0.25,0)"
    ws4.cell(row=r, column=3).value = f"=ROUND(A{r}*0.5,0)"
    ws4.cell(row=r, column=4).value = f"=ROUND(A{r}*0.25,0)"
    ws4.cell(row=r, column=5).value = f"=B{r}*99+C{r}*189.99+D{r}*249.99"
    ws4.cell(row=r, column=5).number_format = money_fmt
    ws4.cell(row=r, column=6).value = f"=E{r}*12"
    ws4.cell(row=r, column=6).number_format = money_fmt
    style_row(ws4, r, 6, r % 2 == 0)

for c in range(1, 7):
    ws4.cell(row=6, column=c).fill = PatternFill("solid", fgColor=LIGHT_GREEN)

for c, w in enumerate([16, 16, 22, 16, 22, 22], 1):
    ws4.column_dimensions[get_column_letter(c)].width = w

# ===== ABA 5: PARAMETROS =====
ws5 = wb.create_sheet("Parametros Sistema")
ws5.sheet_properties.tabColor = AMBER
ws5["A1"] = "PARAMETROS DO SISTEMA \u2014 LEO"
ws5["A1"].font = title_font

h5 = ["Parametro", "Valor", "Unidade", "Observacao"]
for i, h in enumerate(h5, 1):
    ws5.cell(row=3, column=i, value=h)
style_hdr(ws5, 3, 4)

params = [
    ["Periodo Trial", 30, "dias", "Automatico ao cadastrar"],
    ["Franquia Trial", 100, "laudos", "Mesmo do plano Basic"],
    ["Carencia Inadimplencia", 3, "dias", "Apos vencimento, antes de bloquear"],
    ["Rate Limit Emissao", 20, "emissoes/hora", "Protecao anti-fraude"],
    ["Plano Anual (desconto)", 2, "meses gratis", "Futuro - a implementar"],
    ["Excedente Extrato", 1, "gratis/mes/local", "Geracao de extrato financeiro"],
    ["Reemissao sem alteracao", 0, "R$", "Gratis se nao alterar identificacao"],
    ["Gateway Pagamento", "-", "-", "Stripe ou Asaas - a definir"],
]
for r, p in enumerate(params, 4):
    ws5.cell(row=r, column=1, value=p[0])
    ws5.cell(row=r, column=2, value=p[1])
    ws5.cell(row=r, column=3, value=p[2])
    ws5.cell(row=r, column=4, value=p[3])
    ws5.cell(row=r, column=1).alignment = left_align
    ws5.cell(row=r, column=4).alignment = left_align
    if isinstance(p[1], (int, float)):
        ws5.cell(row=r, column=2).font = blue_font
    style_row(ws5, r, 4, r % 2 == 0)

for c, w in enumerate([28, 14, 18, 42], 1):
    ws5.column_dimensions[get_column_letter(c)].width = w

output = "C:\\Users\\sergi\\Desktop\\LEO\\planos_leo.xlsx"
wb.save(output)
print(f"Arquivo salvo em: {output}")
